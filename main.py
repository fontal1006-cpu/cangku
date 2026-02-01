import flet as ft
import openpyxl
import datetime
import os
import shutil
import time

def main(page: ft.Page):
    page.title = "🦅 小鹰提货明细生成器"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.window_width = 380
    page.window_height = 700
    page.scroll = "auto"
    page.padding = 20

    # --- 路径兼容性设置 ---
    # 获取当前脚本所在目录的绝对路径
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 指向 assets 文件夹中的 Excel
    DATA_PATH = os.path.join(base_dir, "assets", "data.xlsx")
    TPL_PATH = os.path.join(base_dir, "assets", "template.xlsx")
    
    # 缓存目录：在手机应用私有目录下创建
    CACHE_DIR = os.path.join(base_dir, "temp_cache")

    # UI 变量
    status_text = ft.Text("", color="blue")
    
    # 初始化检查（防止白屏的关键）
    try:
        if not os.path.exists(CACHE_DIR):
            os.makedirs(CACHE_DIR)
    except Exception as e:
        page.add(ft.Text(f"创建缓存目录失败: {str(e)}", color="red"))

    # UI 控件定义
    search_input = ft.TextField(label="🔍 客户关键字", variant=ft.IndicatorCode.UNDERLINE, border_color="blue")
    product_input = ft.TextField(label="📦 产品名称", variant=ft.IndicatorCode.UNDERLINE)
    count_input = ft.TextField(label="📊 件数", variant=ft.IndicatorCode.UNDERLINE, keyboard_type=ft.KeyboardType.NUMBER)
    temp_dropdown = ft.SegmentedButton(
        segments=[
            ft.Segment(value="常温", label=ft.Text("常温")),
            ft.Segment(value="冷链", label=ft.Text("冷链")),
        ],
        selected={"常温"}
    )

    def clean_cache():
        """清理缓存文件夹"""
        if not os.path.exists(CACHE_DIR):
            return
        for filename in os.listdir(CACHE_DIR):
            file_path = os.path.join(CACHE_DIR, filename)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f"清理失败: {e}")

    def search_customer(keyword):
        # 调试：检查文件是否存在
        if not os.path.exists(DATA_PATH):
            status_text.value = f"❌ 找不到数据库文件:\n{DATA_PATH}"
            page.update()
            return None
        try:
            wb = openpyxl.load_workbook(DATA_PATH, data_only=True)
            ws = wb["Sheet2"]
            matches = {}
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=2).value
                if cell_value and keyword in str(cell_value):
                    info = [
                        cell_value,
                        ws.cell(row=row + 1, column=2).value,
                        ws.cell(row=row + 2, column=2).value,
                        ws.cell(row=row + 3, column=2).value
                    ]
                    matches[str(cell_value)] = info
            wb.close()
            return matches
        except Exception as e:
            status_text.value = f"读取异常: {str(e)}"
            page.update()
            return None

    def handle_generate(e):
        keyword = search_input.value.strip()
        if not keyword:
            page.snack_bar = ft.SnackBar(ft.Text("请输入关键字"))
            page.snack_bar.open = True
            page.update()
            return

        status_text.value = "🔍 正在检索客户..."
        page.update()

        matches = search_customer(keyword)
        if matches is None: return # 报错了
        
        if not matches:
            status_text.value = "❌ 未找到该客户，请检查关键字"
            page.update()
            return

        if len(matches) == 1:
            process_excel(list(matches.values())[0])
        else:
            def select_and_go(name):
                dlg.open = False
                page.update()
                process_excel(matches[name])

            list_items = [ft.ListTile(title=ft.Text(n), on_click=lambda _, n=n: select_and_go(n)) for n in matches.keys()]
            dlg = ft.AlertDialog(title=ft.Text("请选择精确客户"), content=ft.Column(list_items, tight=True))
            page.dialog = dlg
            dlg.open = True
            page.update()

    def process_excel(info):
        try:
            status_text.value = "📝 正在读取模板..."
            page.update()

            if not os.path.exists(TPL_PATH):
                status_text.value = f"❌ 找不到模板文件:\n{TPL_PATH}"
                page.update()
                return

            # 打开模板
            wb = openpyxl.load_workbook(TPL_PATH)
            ws = wb["1"]

            # 填写数据
            today = datetime.datetime.now()
            ws["C2"] = today.strftime("%Y年%m月%d日")
            ws["B6"], ws["E6"], ws["C6"], ws["D6"] = info[0], info[1], info[2], info[3]
            ws["G6"] = product_input.value
            ws["J6"] = count_input.value
            ws["M6"] = list(temp_dropdown.selected)[0]

            # 保存到临时缓存
            filename = f"提货明细_{info[0]}_{today.strftime('%m%d%H%M')}.xlsx"
            temp_file_path = os.path.join(CACHE_DIR, filename)
            wb.save(temp_file_path)
            wb.close()

            status_text.value = f"✅ 生成成功！正在唤起分享..."
            page.update()

            # 唤起手机分享（Flet 在安卓上的高级功能）
            page.share_files([temp_file_path])
            
        except Exception as ex:
            status_text.value = f"🚨 程序错误: {str(ex)}"
            page.update()

    # --- UI 布局 ---
    page.add(
        ft.Column([
            ft.Container(
                content=ft.Text("🦅 小鹰提货生成器", size=28, weight="bold", color="blue"),
                alignment=ft.alignment.center,
                padding=20
            ),
            search_input,
            product_input,
            count_input,
            ft.Text("🌡️ 选择温度:"),
            temp_dropdown,
            ft.Divider(height=10, color="transparent"),
            ft.ElevatedButton(
                "🚀 生成并打开分享",
                on_click=handle_generate,
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10)),
                width=400,
                height=50
            ),
            ft.Container(status_text, alignment=ft.alignment.center, padding=10)
        ])
    )

# 确保 assets_dir 指向正确的文件夹名称
ft.app(target=main, assets_dir="assets")
